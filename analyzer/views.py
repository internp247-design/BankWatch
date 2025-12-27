import os
from django.conf import settings
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import models
from django.contrib import messages
from django.db import transaction as db_transaction
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from datetime import timedelta
from django.utils import timezone

from .models import BankAccount, BankStatement, Transaction, AnalysisSummary, Rule, CustomCategory, CustomCategoryRule, CustomCategoryRuleCondition
from .forms import BankStatementForm
from .rules_forms import RuleForm, RuleConditionFormSet, CustomCategoryForm, CustomCategoryRuleForm, CustomCategoryRuleConditionFormSet
from .rules_engine import RulesEngine, categorize_with_rules
from collections import defaultdict

# Import file parsers with error handling
try:
    from .file_parsers import StatementParser
    FILE_PARSERS_AVAILABLE = True
except ImportError:
    FILE_PARSERS_AVAILABLE = False
    print("Warning: File parsers not available. Install required packages.")

try:
    from .pdf_parser import categorize_transaction
    CATEGORY_PARSER_AVAILABLE = True
except ImportError:
    CATEGORY_PARSER_AVAILABLE = False
    print("Warning: Category parser not available.")

# Import Excel export dependencies
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    EXCEL_EXPORT_AVAILABLE = False
    print("Warning: openpyxl not available. Excel export will be disabled.")

@login_required
def dashboard(request):
    accounts = BankAccount.objects.filter(user=request.user)
    
    # If user has no accounts, create a default one
    if not accounts.exists():
        BankAccount.objects.create(
            user=request.user,
            account_name="Primary Account",
            bank_name="My Bank"
        )
        accounts = BankAccount.objects.filter(user=request.user)
    
    # Get recent transactions and analysis data
    recent_transactions = Transaction.objects.filter(
        statement__account__user=request.user
    ).select_related('statement').order_by('-date')[:10]
    
    # Calculate summary data from all statements
    all_transactions = Transaction.objects.filter(
        statement__account__user=request.user
    )
    
    total_income = all_transactions.filter(transaction_type='CREDIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    total_expenses = all_transactions.filter(transaction_type='DEBIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    net_savings = total_income - total_expenses
    
    # Get recent statements
    recent_statements = BankStatement.objects.filter(
        account__user=request.user
    ).order_by('-upload_date')[:5]
    
    # Calculate category totals for charts
    category_totals = {}
    expense_transactions = all_transactions.filter(transaction_type='DEBIT')
    
    for category_code, category_name in Transaction.CATEGORY_CHOICES:
        if category_code != 'INCOME':
            category_total = expense_transactions.filter(category=category_code).aggregate(
                total=models.Sum('amount')
            )['total'] or 0
            if category_total > 0:
                category_totals[category_code] = {
                    'name': category_name,
                    'amount': float(category_total),
                    'percentage': 0
                }
    
    # Calculate percentages
    total_expenses_amount = sum(cat['amount'] for cat in category_totals.values())
    for category in category_totals.values():
        if total_expenses_amount > 0:
            category['percentage'] = round((category['amount'] / total_expenses_amount) * 100, 1)
    
    # Financial health score
    if total_income > 0:
        savings_rate = (net_savings / total_income) * 100
        if savings_rate >= 20:
            financial_health = {'score': 85, 'status': 'Excellent', 'message': 'Great savings rate!'}
        elif savings_rate >= 10:
            financial_health = {'score': 70, 'status': 'Good', 'message': 'Solid financial health'}
        else:
            financial_health = {'score': 50, 'status': 'Needs Attention', 'message': 'Consider increasing savings'}
    else:
        financial_health = {'score': 0, 'status': 'No Data', 'message': 'Upload statements to see your financial health'}
    
    context = {
        'accounts': accounts,
        'recent_transactions': recent_transactions,
        'monthly_income': total_income,
        'monthly_expenses': total_expenses,
        'monthly_savings': net_savings,
        'category_totals': category_totals,
        'financial_health': financial_health,
        'recent_statements': recent_statements,
    }
    
    return render(request, 'analyzer/dashboard.html', context)

@login_required
def upload_statement(request):
    """Handle upload of PDF, Excel, and CSV statements"""
    if request.method == 'POST':
        form = BankStatementForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                # Get the account from form
                account = form.cleaned_data['account']
                
                # Check if user owns this account
                if account.user != request.user:
                    messages.error(request, 'You do not have permission to upload to this account.')
                    return redirect('upload_statement')
                
                # Get the filename
                filename = request.FILES['statement_file'].name
                
                # Check for duplicate statement on the same account
                existing_statements = BankStatement.objects.filter(
                    account=account,
                    original_filename=filename
                )
                
                if existing_statements.exists():
                    messages.warning(
                        request,
                        f'⚠️ Warning: A statement with the filename "{filename}" has already been uploaded to this account. '
                        f'Are you sure you want to upload it again? This may create duplicate transactions.'
                    )
                    return redirect('upload_statement')
                
                # Save the statement
                statement = form.save(commit=False)
                statement.original_filename = filename
                
                # Determine file type from filename
                filename = request.FILES['statement_file'].name.lower()
                if filename.endswith('.pdf'):
                    statement.file_type = BankStatement.PDF
                elif filename.endswith(('.xlsx', '.xls')):
                    statement.file_type = BankStatement.EXCEL
                elif filename.endswith('.csv'):
                    statement.file_type = BankStatement.CSV
                
                statement.save()
                
                if not FILE_PARSERS_AVAILABLE:
                    messages.error(request, 
                        'File parsing libraries not installed. '
                        'Please install: pip install pandas openpyxl xlrd pdfplumber'
                    )
                    return redirect('upload_statement')
                
                # Process the uploaded file
                file_path = os.path.join(settings.MEDIA_ROOT, str(statement.statement_file))
                
                try:
                    # Extract transactions
                    transactions_data = StatementParser.parse_file(file_path, statement.file_type)
                    
                    print(f"Extracted {len(transactions_data)} transactions from {statement.file_type} file")
                    
                    # Import UPI parser for better categorization
                    try:
                        from .upi_parser import UPIParser
                        UPI_PARSER_AVAILABLE = True
                    except ImportError:
                        UPI_PARSER_AVAILABLE = False
                    
                    # Create Transaction objects
                    created_count = 0
                    for transaction_data in transactions_data:
                        # Determine category with multiple strategies
                        category = 'OTHER'
                        desc = transaction_data['description'].upper()
                        
                        # Strategy 1: UPI parsing if available
                        if UPI_PARSER_AVAILABLE and UPIParser.is_upi_description(transaction_data['description']):
                            upi_data = UPIParser.parse_upi_fields(transaction_data['description'])
                            if 'purpose' in upi_data:
                                purpose = upi_data['purpose'].upper()
                            else:
                                purpose = desc
                            
                            # Check for keywords in both description and purpose
                            full_text = desc + ' ' + purpose
                            
                            if any(word in full_text for word in ['PURCHASE', 'SHOPPING', 'STORE', 'AMAZON', 'FLIPKART']):
                                category = 'SHOPPING'
                            elif any(word in full_text for word in ['FOOD', 'RESTAURANT', 'PIZZA', 'CAFE', 'ZOMATO']):
                                category = 'FOOD'
                            elif any(word in full_text for word in ['TAXI', 'TRANSPORT', 'METRO', 'CARZONRENT', 'PETROL']):
                                category = 'TRANSPORT'
                            elif any(word in full_text for word in ['MEDICINE', 'HEALTH', 'DENTAL', 'DOCTOR']):
                                category = 'HEALTHCARE'
                            elif any(word in full_text for word in ['TRAVEL', 'HOTEL', 'BOOKING', 'FLIGHT']):
                                category = 'TRAVEL'
                            elif any(word in full_text for word in ['ENTERTAINMENT', 'MOVIE', 'CINEMA']):
                                category = 'ENTERTAINMENT'
                            elif any(word in full_text for word in ['BILL', 'ELECTRICITY', 'INTERNET', 'AIRTEL']):
                                category = 'BILLS'
                        
                        # Strategy 2: Use rules engine if categorization not done
                        if category == 'OTHER':
                            rules_category = categorize_with_rules(transaction_data, request.user)
                            if rules_category:
                                category = rules_category
                        
                        # Strategy 3: Use ML parser if available and still OTHER
                        if category == 'OTHER' and CATEGORY_PARSER_AVAILABLE:
                            ml_category = categorize_transaction(
                                transaction_data['description'],
                                transaction_data['amount'],
                                transaction_data['transaction_type']
                            )
                            if ml_category:
                                category = ml_category
                        
                        Transaction.objects.create(
                            statement=statement,
                            date=transaction_data['date'],
                            description=transaction_data['description'][:500],
                            amount=transaction_data['amount'],
                            transaction_type=transaction_data['transaction_type'],
                            category=category
                        )
                        created_count += 1
                    
                    # Create Analysis Summary
                    transactions = Transaction.objects.filter(statement=statement)
                    total_income = sum(t.amount for t in transactions if t.transaction_type == 'CREDIT')
                    total_expenses = sum(t.amount for t in transactions if t.transaction_type == 'DEBIT')
                    
                    AnalysisSummary.objects.create(
                        statement=statement,
                        total_income=total_income,
                        total_expenses=total_expenses,
                        net_savings=total_income - total_expenses
                    )
                    
                    messages.success(request, 
                        f'✅ Successfully uploaded and analyzed {statement.get_file_type_display()} file! '
                        f'Found {created_count} transactions.'
                    )
                    return redirect('statement_rules_prompt', statement_id=statement.id)
                    
                except Exception as e:
                    # If file processing fails, delete the statement and show error
                    statement.delete()
                    error_msg = f'Error processing file: {str(e)}'
                    print(error_msg)
                    messages.error(request, error_msg)
                    return render(request, 'analyzer/upload.html', {'form': form})
                    
            except Exception as e:
                messages.error(request, f'Error saving statement: {str(e)}')
                return render(request, 'analyzer/upload.html', {'form': form})
    else:
        form = BankStatementForm()
        form.fields['account'].queryset = BankAccount.objects.filter(user=request.user)
    
    return render(request, 'analyzer/upload.html', {'form': form})

@login_required
def analysis_results(request, statement_id):
    statement = get_object_or_404(BankStatement, id=statement_id, account__user=request.user)
    transactions = Transaction.objects.filter(statement=statement).order_by('-date')
    
    # Calculate summary data
    total_income = sum(t.amount for t in transactions if t.transaction_type == 'CREDIT')
    total_expenses = sum(t.amount for t in transactions if t.transaction_type == 'DEBIT')
    net_savings = total_income - total_expenses
    
    # Calculate category totals for the chart
    category_totals = {}
    for transaction in transactions.filter(transaction_type='DEBIT'):
        category = transaction.get_category_display()
        category_totals[category] = category_totals.get(category, 0) + float(transaction.amount)
    
    # Comprehensive color palette - unique colors for each category
    color_palette = [
        '#e67e22',  # Orange - Food
        '#9b59b6',  # Purple - Shopping
        '#3498db',  # Blue - Transport
        '#e74c3c',  # Red - Bills
        '#f1c40f',  # Yellow - Entertainment
        '#1abc9c',  # Turquoise
        '#34495e',  # Dark Gray
        '#16a085',  # Green
        '#d35400',  # Dark Orange
        '#8e44ad',  # Dark Purple
        '#c0392b',  # Dark Red
        '#27ae60',  # Green
        '#2980b9',  # Dark Blue
        '#f39c12',  # Golden
        '#95a5a6',  # Gray
    ]
    
    chart_labels = list(category_totals.keys())
    chart_data = list(category_totals.values())
    # Assign unique colors to each category in order
    chart_colors = [color_palette[i % len(color_palette)] for i in range(len(chart_labels))]
    
    # Get analysis summary if exists
    try:
        analysis_summary = AnalysisSummary.objects.get(statement=statement)
    except AnalysisSummary.DoesNotExist:
        analysis_summary = None
    
    # Get user's custom categories
    custom_categories = CustomCategory.objects.filter(user=request.user, is_active=True)
    
    context = {
        'statement': statement,
        'transactions': transactions,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_savings': net_savings,
        'category_totals': category_totals,
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'chart_colors': chart_colors,
        'analysis_summary': analysis_summary,
        'custom_categories': custom_categories,
    }
    return render(request, 'analyzer/results.html', context)

@login_required
def create_first_account(request):
    if request.method == 'POST':
        account_name = request.POST.get('account_name')
        bank_name = request.POST.get('bank_name')
        
        if account_name:
            account = BankAccount.objects.create(
                user=request.user,
                account_name=account_name,
                bank_name=bank_name or 'My Bank'
            )
            messages.success(request, f'Account "{account_name}" created successfully!')
            return redirect('upload_statement')
    
    return render(request, 'analyzer/create_account.html')


@login_required
def create_account(request):
    """Create a new bank account for the logged-in user (general use)."""
    if request.method == 'POST':
        bank_name = request.POST.get('bank_name')
        account_name = request.POST.get('account_name')
        account_type = request.POST.get('account_type')
        account_number = request.POST.get('account_number')
        ifsc_code = request.POST.get('ifsc_code')
        description = request.POST.get('description')

        if bank_name:
            BankAccount.objects.create(
                user=request.user,
                bank_name=bank_name,
                account_name=account_name or bank_name,
                account_type=account_type,
                account_number=account_number,
                ifsc_code=ifsc_code,
                description=description
            )
            messages.success(request, f'Account "{bank_name}" created successfully!')
            return redirect('dashboard')

    return render(request, 'analyzer/create_account.html')

# ===========================================
# RULES ENGINE VIEWS
# ===========================================

@login_required
def rules_list(request):
    """List all rules"""
    rules = Rule.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'analyzer/rules_list.html', {'rules': rules})

@login_required
def create_rule(request):
    """Create a new rule"""
    if request.method == 'POST':
        form = RuleForm(request.POST)
        if form.is_valid():
            rule = form.save(commit=False)
            rule.user = request.user
            rule.save()
            messages.success(request, f'Rule "{rule.name}" created successfully!')
            return redirect('edit_rule', rule_id=rule.id)
    else:
        form = RuleForm()
    
    return render(request, 'analyzer/create_rule.html', {'form': form})

@login_required
def edit_rule(request, rule_id):
    """Edit an existing rule"""
    rule = get_object_or_404(Rule, id=rule_id, user=request.user)
    
    if request.method == 'POST':
        form = RuleForm(request.POST, instance=rule)
        formset = RuleConditionFormSet(request.POST, instance=rule)
        
        if form.is_valid() and formset.is_valid():
            try:
                form.save()
                formset.save()
                messages.success(request, f'Rule "{rule.name}" updated successfully!')
                return redirect('rules_list')
            except Exception as e:
                messages.error(request, f'Error saving rule: {str(e)}')
        else:
            # Form validation failed
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f"{field}: {error}")
            for formset_error in formset.non_form_errors():
                error_messages.append(formset_error)
            
            if error_messages:
                messages.error(request, 'Please correct the errors: ' + '; '.join(error_messages))
    else:
        form = RuleForm(instance=rule)
        formset = RuleConditionFormSet(instance=rule)
    
    return render(request, 'analyzer/edit_rule.html', {
        'form': form,
        'formset': formset,
        'rule': rule
    })

@login_required
def delete_rule(request, rule_id):
    """Delete a rule"""
    rule = get_object_or_404(Rule, id=rule_id, user=request.user)
    if request.method == 'POST':
        rule_name = rule.name
        rule.delete()
        messages.success(request, f'Rule "{rule_name}" deleted successfully!')
        return redirect('rules_list')
    
    return render(request, 'analyzer/delete_rule.html', {'rule': rule})

@login_required
def apply_rules(request):
    """Apply rules to existing transactions"""
    try:
        # Get stats for display
        active_rules_count = Rule.objects.filter(user=request.user, is_active=True).count()
        total_transactions = Transaction.objects.filter(
            statement__account__user=request.user
        ).count()
        
        if request.method == 'POST':
            # Support account-scoped application: an optional account_id can be passed
            account_id = request.POST.get('account_id') or request.GET.get('account_id')

            # Support AJAX requests so the frontend can stop the spinner and show results
            engine = RulesEngine(request.user)
            if account_id:
                transactions = Transaction.objects.filter(
                    statement__account__user=request.user,
                    statement__account_id=account_id
                )
            else:
                transactions = Transaction.objects.filter(
                    statement__account__user=request.user
                )

            updated_count = 0
            updated_ids = []
            prev_map = {}
            matched_map = {}
            with db_transaction.atomic():
                for transaction in transactions:
                    transaction_data = {
                        'date': transaction.date,
                        'description': transaction.description,
                        'amount': float(transaction.amount),
                        'transaction_type': transaction.transaction_type
                    }

                    # Determine which rule (if any) matches and the target category
                    matched_rule = engine.find_matching_rule(transaction_data)
                    category = matched_rule.category if matched_rule else None

                    if category and category != transaction.category:
                        # record previous category so we can show changes
                        prev_map[str(transaction.id)] = transaction.category
                        # record which rule matched
                        matched_map[str(transaction.id)] = matched_rule.name if matched_rule else None
                        transaction.category = category
                        transaction.save()
                        updated_count += 1
                        updated_ids.append(transaction.id)

            # If AJAX request, return JSON so JS can stop spinner and update UI
            # Store list of updated ids and previous categories in session so results view can show only changed
            if updated_ids:
                request.session['last_rules_applied_ids'] = updated_ids
                request.session['last_rules_applied_prev'] = prev_map
                request.session['last_rules_applied_rule'] = matched_map
            else:
                # clear previous session keys if nothing changed
                request.session.pop('last_rules_applied_ids', None)
                request.session.pop('last_rules_applied_prev', None)
                request.session.pop('last_rules_applied_rule', None)

            if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
                redirect_url = reverse('rules_application_results')
                sep = '?'
                if account_id:
                    redirect_url += f'?account_id={account_id}'
                    sep = '&'
                # instruct results page to show only changed transactions
                redirect_url += f"{sep}show_changed=1"
                return JsonResponse({
                    'status': 'ok',
                    'updated': updated_count,
                    'total': transactions.count(),
                    'message': f'Rules applied successfully! Updated {updated_count} out of {transactions.count()} transactions.',
                    'redirect_url': redirect_url,
                })

            messages.success(request,
                f'Rules applied successfully! Updated {updated_count} out of {transactions.count()} transactions.'
            )
            return redirect('rules_list')
        
        # Provide accounts for the selector
        accounts = BankAccount.objects.filter(user=request.user)

        return render(request, 'analyzer/apply_rules.html', {
            'active_rules_count': active_rules_count,
            'total_transactions': total_transactions,
            'accounts': accounts,
        })
    except Exception as e:
        import traceback
        error_msg = f"ERROR in apply_rules: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
            return JsonResponse({
                'status': 'error',
                'message': f'Error applying rules: {str(e)}'
            }, status=500)
        messages.error(request, f"Error applying rules: {str(e)}")
        return redirect('rules_list')

@login_required
def test_rules(request):
    """Test rules against sample transactions"""
    if request.method == 'POST':
        test_description = request.POST.get('description', '').strip()
        test_amount_str = request.POST.get('amount', '0')
        test_date = request.POST.get('date', '')
        
        # Validate amount
        try:
            test_amount = float(test_amount_str)
        except ValueError:
            test_amount = 0
        
        transaction_data = {
            'description': test_description,
            'amount': test_amount,
            'date': test_date,
            'transaction_type': 'DEBIT' if test_amount < 0 else 'CREDIT'
        }
        
        engine = RulesEngine(request.user)
        matching_rules = []
        
        for rule in engine.rules:
            if engine._matches_rule(transaction_data, rule):
                matching_rules.append(rule)
        
        return render(request, 'analyzer/test_rules.html', {
            'test_description': test_description,
            'test_amount': test_amount,
            'test_date': test_date,
            'matching_rules': matching_rules,
            'transaction_data': transaction_data
        })
    
    return render(request, 'analyzer/test_rules.html')


@login_required
def rules_application_results(request):
    """Show which rule (if any) and custom category (if any) matches each transaction. Supports optional account filter and filtering by selected rules/categories."""
    try:
        account_id = request.GET.get('account_id')
        show_changed = request.GET.get('show_changed') in ['1', 'true', 'True']
        
        print(f"DEBUG: rules_application_results called with show_changed={show_changed}, account_id={account_id}")
        
        # Get selected rule IDs and category IDs from GET parameters
        selected_rule_ids = request.GET.getlist('rule_ids')
        selected_category_ids = request.GET.getlist('category_ids')
        
        # Convert to integers
        selected_rule_ids = [int(rid) for rid in selected_rule_ids if rid.isdigit()]
        selected_category_ids = [int(cid) for cid in selected_category_ids if cid.isdigit()]
        
        print(f"DEBUG: selected_rule_ids={selected_rule_ids}, selected_category_ids={selected_category_ids}")
        
        engine = RulesEngine(request.user)
        
        # Import custom category engine
        from .rules_engine import CustomCategoryRulesEngine
        custom_category_engine = CustomCategoryRulesEngine(request.user)
        
        # If show_changed requested, read the list of ids from session (set by apply_rules)
        if show_changed:
            updated_ids = request.session.get('last_rules_applied_ids', [])
            print(f"DEBUG: show_changed=True, updated_ids={updated_ids}")
            # ensure we only fetch transactions belonging to this user (and optional account)
            if account_id:
                transactions = Transaction.objects.filter(
                    id__in=updated_ids,
                    statement__account__user=request.user,
                    statement__account_id=account_id
                ).select_related('statement', 'statement__account').order_by('-date')
            else:
                transactions = Transaction.objects.filter(
                    id__in=updated_ids,
                    statement__account__user=request.user
                ).select_related('statement', 'statement__account').order_by('-date')
        else:
            if account_id:
                transactions = Transaction.objects.filter(
                    statement__account__user=request.user,
                    statement__account_id=account_id
                ).select_related('statement', 'statement__account').order_by('-date')
            else:
                transactions = Transaction.objects.filter(
                    statement__account__user=request.user
                ).select_related('statement', 'statement__account').order_by('-date')

        results = []
        for tx in transactions:
            try:
                tx_data = {
                    'date': tx.date,
                    'description': tx.description,
                    'amount': float(tx.amount),
                    'transaction_type': tx.transaction_type
                }
                
                # Check for rule match
                matched_rule = engine.find_matching_rule(tx_data)
                matched_rule_id = matched_rule.id if matched_rule else None
                matched_rule_category = matched_rule.category if matched_rule else None
                matched_rule_name = matched_rule.name if matched_rule else None
                
                # Check for custom category match
                matched_custom_category = custom_category_engine.apply_rules_to_transaction(tx_data)
                matched_custom_category_id = matched_custom_category.id if matched_custom_category else None
                matched_custom_category_name = matched_custom_category.name if matched_custom_category else None
                
                # Only include if there's a match (rule or custom category)
                if matched_rule_name or matched_custom_category_name:
                    # Safely access statement and account
                    account_name = 'Unknown'
                    account_id = None
                    try:
                        if tx.statement and tx.statement.account:
                            account_name = tx.statement.account.account_name
                            account_id = tx.statement.account.id
                    except (AttributeError, ObjectDoesNotExist):
                        pass
                    
                    results.append({
                        'id': tx.id,
                        'date': str(tx.date),  # Convert date to string for session serialization
                        'description': tx.description,
                        'amount': float(tx.amount),
                        'current_category': tx.category,
                        'matched_rule_id': matched_rule_id,
                        'matched_rule_category': matched_rule_category,
                        'matched_rule_name': matched_rule_name,
                        'matched_custom_category_id': matched_custom_category_id,
                        'matched_custom_category_name': matched_custom_category_name,
                        'previous_category': request.session.get('last_rules_applied_prev', {}).get(str(tx.id)) if show_changed else None,
                        'account_name': account_name,
                        'account_id': account_id,
                    })
            except Exception as e:
                print(f"ERROR - Failed to process transaction {tx.id}: {str(e)}")
                continue

        # Get list of user's accounts for selector
        accounts = BankAccount.objects.filter(user=request.user)
        no_changes = show_changed and len(results) == 0

        # Get all user's custom categories and rules for filter panels
        all_custom_categories = CustomCategory.objects.filter(user=request.user, is_active=True).order_by('name')
        all_rules = Rule.objects.filter(user=request.user, is_active=True).order_by('name')
        
        # Filter categories and rules based on selection
        if selected_category_ids:
            custom_categories = all_custom_categories.filter(id__in=selected_category_ids)
        else:
            custom_categories = CustomCategory.objects.none()
        
        if selected_rule_ids:
            rules = all_rules.filter(id__in=selected_rule_ids)
        else:
            rules = Rule.objects.none()

        # Compute summary report table - ONLY for selected rules and categories
        rule_category_report = []
        
        if selected_rule_ids or selected_category_ids:
            # Initialize data for selected items
            if selected_rule_ids:
                for rule in rules:
                    rule_category_report.append({
                        'type': 'rule',
                        'id': rule.id,
                        'name': rule.name,
                        'category': rule.get_category_display(),
                        'transaction_count': 0,
                        'total_amount': 0.0
                    })
            
            if selected_category_ids:
                for category in custom_categories:
                    rule_category_report.append({
                        'type': 'category',
                        'id': category.id,
                        'name': category.name,
                        'category': 'Custom',
                        'transaction_count': 0,
                        'total_amount': 0.0
                    })
            
            # Populate counts and amounts from results
            for result in results:
                # Update rule if matched and selected
                if result['matched_rule_id'] and result['matched_rule_id'] in selected_rule_ids:
                    for item in rule_category_report:
                        if item['type'] == 'rule' and item['id'] == result['matched_rule_id']:
                            item['transaction_count'] += 1
                            item['total_amount'] += float(result['amount'] or 0)
                
                # Update category if matched and selected
                if result['matched_custom_category_id'] and result['matched_custom_category_id'] in selected_category_ids:
                    for item in rule_category_report:
                        if item['type'] == 'category' and item['id'] == result['matched_custom_category_id']:
                            item['transaction_count'] += 1
                            item['total_amount'] += float(result['amount'] or 0)
        
        # Filter results to show only those matching selected rules/categories
        filtered_results = []
        if selected_rule_ids or selected_category_ids:
            for r in results:
                include = False
                # Include if matched rule is selected
                if r['matched_rule_id'] and r['matched_rule_id'] in selected_rule_ids:
                    include = True
                # Include if matched category is selected
                if r['matched_custom_category_id'] and r['matched_custom_category_id'] in selected_category_ids:
                    include = True
                
                if include:
                    filtered_results.append(r)
        else:
            filtered_results = results

        # Store filtered results in session for export functions
        request.session['export_filtered_results'] = filtered_results
        request.session['export_selected_rule_ids'] = selected_rule_ids
        request.session['export_selected_category_ids'] = selected_category_ids
        request.session.modified = True  # Ensure session is saved

        # compute colspan for template (base 7 columns + previous column if show_changed)
        colspan = 7 + (1 if show_changed else 0)

        return render(request, 'analyzer/apply_rules_results.html', {
            'results': filtered_results,
            'all_results': results,
            'rule_category_report': rule_category_report,
            'accounts': accounts,
            'selected_account_id': int(account_id) if account_id else None,
            'show_changed': show_changed,
            'no_changes': show_changed and len(filtered_results) == 0,
            'colspan': colspan,
            'custom_categories': all_custom_categories,
            'all_custom_categories': all_custom_categories,
            'rules': all_rules,
            'selected_rule_ids': selected_rule_ids,
            'selected_category_ids': selected_category_ids,
        })
        
    except Exception as e:
        import traceback
        error_msg = f"ERROR in rules_application_results: {str(e)}\n{traceback.format_exc()}"
        print(error_msg)
        messages.error(request, f"Error loading results: {str(e)}")
        return redirect('rules_list')



@login_required
def rules_categorized(request):
    """Show rules grouped by category for the current user."""
    user_rules = Rule.objects.filter(user=request.user).order_by('category', '-created_at')
    grouped = defaultdict(list)
    for rule in user_rules:
        grouped[rule.get_category_display()].append(rule)

    # Convert to list of tuples for template ordering
    grouped_list = [(category, grouped[category]) for category in sorted(grouped.keys())]

    return render(request, 'analyzer/rules_categorized.html', {
        'grouped_rules': grouped_list,
    })

@login_required
def bulk_delete_transactions(request, statement_id=None):
    """Bulk delete transactions (for cleanup)"""
    if request.method == 'POST':
        if statement_id:
            # Delete transactions for specific statement
            transactions = Transaction.objects.filter(
                statement_id=statement_id,
                statement__account__user=request.user
            )
            statement = get_object_or_404(BankStatement, id=statement_id, account__user=request.user)
            count = transactions.count()
            transactions.delete()
            messages.success(request, f'Deleted {count} transactions from statement.')
            return redirect('dashboard')
        else:
            # Delete all user's transactions
            transactions = Transaction.objects.filter(
                statement__account__user=request.user
            )
            count = transactions.count()
            transactions.delete()
            messages.success(request, f'Deleted all {count} transactions.')
            return redirect('dashboard')
    
    if statement_id:
        statement = get_object_or_404(BankStatement, id=statement_id, account__user=request.user)
        return render(request, 'analyzer/confirm_delete.html', {
            'statement': statement,
            'object_type': 'transactions'
        })
    else:
        return render(request, 'analyzer/confirm_delete.html', {
            'object_type': 'all transactions'
        })

@login_required
def statement_rules_prompt(request, statement_id):
    """Show prompt to apply global rules to a statement"""
    statement = get_object_or_404(BankStatement, id=statement_id, account__user=request.user)
    
    # Check if statement already has rules applied
    if statement.rules_applied:
        return redirect('analysis_results', statement_id=statement_id)
    
    # Count active rules
    active_rules_count = Rule.objects.filter(user=request.user, is_active=True).count()
    
    if request.method == 'POST':
        apply_rules_choice = request.POST.get('apply_rules')
        
        if apply_rules_choice == 'yes':
            # Apply global rules to this statement
            transactions = Transaction.objects.filter(statement=statement)
            engine = RulesEngine(request.user)
            updated_count = 0
            
            with db_transaction.atomic():
                for transaction in transactions:
                    transaction_data = {
                        'date': transaction.date,
                        'description': transaction.description,
                        'amount': float(transaction.amount),
                        'transaction_type': transaction.transaction_type
                    }
                    
                    matched_rule = engine.find_matching_rule(transaction_data)
                    if matched_rule and matched_rule.category != transaction.category:
                        transaction.category = matched_rule.category
                        transaction.save()
                        updated_count += 1
            
            statement.rules_applied = True
            statement.save()
            
            messages.success(request, 
                f'✅ Applied global rules! Updated {updated_count} transaction(s).')
        else:
            # Mark as reviewed without applying rules
            statement.rules_applied = True
            statement.save()
            messages.info(request, 'Proceeding without applying global rules.')
        
        return redirect('analysis_results', statement_id=statement_id)
    
    context = {
        'statement': statement,
        'active_rules_count': active_rules_count,
    }
    return render(request, 'analyzer/statement_rules_prompt.html', context)

@login_required
def toggle_rule_active(request):
    """AJAX endpoint to toggle rule active status"""
    if request.method == 'POST' and request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        rule_id = request.POST.get('rule_id')
        try:
            rule = Rule.objects.get(id=rule_id, user=request.user)
            rule.is_active = not rule.is_active
            rule.save()
            return JsonResponse({
                'success': True,
                'is_active': rule.is_active,
                'message': f'Rule "{rule.name}" is now {"active" if rule.is_active else "inactive"}.'
            })
        except Rule.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Rule not found.'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)


@login_required
def view_account_details(request, account_id):
    """View account-specific financial overview"""
    account = get_object_or_404(BankAccount, id=account_id, user=request.user)
    
    # Get transactions for this account only
    account_transactions = Transaction.objects.filter(
        statement__account=account
    ).select_related('statement').order_by('-date')
    
    recent_transactions = account_transactions[:15]
    
    # Calculate summary data for this account only
    total_income = account_transactions.filter(transaction_type='CREDIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    total_expenses = account_transactions.filter(transaction_type='DEBIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    net_savings = total_income - total_expenses
    
    # Calculate category totals for charts
    category_totals = {}
    expense_transactions = account_transactions.filter(transaction_type='DEBIT')
    
    for category_code, category_name in Transaction.CATEGORY_CHOICES:
        if category_code != 'INCOME':
            category_total = expense_transactions.filter(category=category_code).aggregate(
                total=models.Sum('amount')
            )['total'] or 0
            if category_total > 0:
                category_totals[category_code] = {
                    'name': category_name,
                    'amount': float(category_total),
                    'percentage': 0
                }
    
    # Calculate percentages
    total_expenses_amount = sum(cat['amount'] for cat in category_totals.values())
    for category in category_totals.values():
        if total_expenses_amount > 0:
            category['percentage'] = round((category['amount'] / total_expenses_amount) * 100, 1)
    
    # Financial health score
    if total_income > 0:
        savings_rate = (net_savings / total_income) * 100
        if savings_rate >= 20:
            financial_health = {'score': 85, 'status': 'Excellent', 'message': 'Great savings rate!'}
        elif savings_rate >= 10:
            financial_health = {'score': 70, 'status': 'Good', 'message': 'Solid financial health'}
        else:
            financial_health = {'score': 50, 'status': 'Needs Attention', 'message': 'Consider increasing savings'}
    else:
        financial_health = {'score': 0, 'status': 'No Data', 'message': 'No transactions for this account'}
    
    context = {
        'account': account,
        'transactions': account_transactions,
        'recent_transactions': recent_transactions,
        'monthly_income': total_income,
        'monthly_expenses': total_expenses,
        'monthly_savings': net_savings,
        'category_totals': category_totals,
        'financial_health': financial_health,
    }
    
    return render(request, 'analyzer/account_details.html', context)


@login_required
def delete_account(request, account_id):
    """Delete a bank account"""
    account = get_object_or_404(BankAccount, id=account_id, user=request.user)
    
    if request.method == 'POST':
        account_name = account.account_name
        account.delete()
        messages.success(request, f'Account "{account_name}" has been deleted successfully.')
        return redirect('dashboard')
    
    # Show confirmation page for GET
    context = {'account': account}
    return render(request, 'analyzer/delete_account.html', context)


@login_required
def change_rule_status_on_results(request):
    """Toggle rule active status from results page (AJAX endpoint)"""
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        rule_id = request.POST.get('rule_id')
        
        try:
            rule = Rule.objects.get(id=rule_id)
            rule.is_active = not rule.is_active
            rule.save()
            return JsonResponse({
                'success': True,
                'is_active': rule.is_active,
                'message': f'Rule is now {"active" if rule.is_active else "inactive"}.'
            })
        except Rule.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Rule not found.'}, status=404)
    
    return JsonResponse({'success': False, 'error': 'Invalid request.'}, status=400)


# ============= CUSTOM CATEGORY VIEWS =============

@login_required
def create_custom_category_and_rule(request):
    """Create a custom category and its rule in one interface"""
    if request.method == 'POST':
        category_form = CustomCategoryForm(request.POST)
        
        if category_form.is_valid():
            # Save custom category
            custom_category = category_form.save(commit=False)
            custom_category.user = request.user
            custom_category.save()
            
            # Redirect to create rule for this category
            messages.success(request, f'Custom category "{custom_category.name}" created successfully!')
            return redirect('create_custom_category_rule', category_id=custom_category.id)
    else:
        category_form = CustomCategoryForm()
    
    context = {
        'form': category_form,
        'page_title': 'Create Custom Category',
        'step': 1,
        'step_title': 'Step 1: Create Custom Category'
    }
    return render(request, 'analyzer/create_custom_category.html', context)


@login_required
def create_custom_category_rule(request, category_id):
    """Create a rule for a custom category"""
    custom_category = get_object_or_404(CustomCategory, id=category_id, user=request.user)
    
    if request.method == 'POST':
        rule_form = CustomCategoryRuleForm(request.POST)
        
        if rule_form.is_valid():
            with db_transaction.atomic():
                # Save rule first
                rule = rule_form.save(commit=False)
                rule.user = request.user
                rule.custom_category = custom_category
                rule.save()
                
                # Now bind the formset to the saved rule
                condition_formset = CustomCategoryRuleConditionFormSet(request.POST, instance=rule)
                
                if condition_formset.is_valid():
                    # Save conditions
                    condition_formset.save()
                    messages.success(request, f'Rule "{rule.name}" created successfully for "{custom_category.name}"!')
                    return redirect('custom_categories_list')
                else:
                    # If formset is invalid, delete the rule and redisplay the form with errors
                    rule.delete()
        else:
            condition_formset = CustomCategoryRuleConditionFormSet()
    else:
        rule_form = CustomCategoryRuleForm()
        condition_formset = CustomCategoryRuleConditionFormSet()
    
    context = {
        'custom_category': custom_category,
        'rule_form': rule_form,
        'condition_formset': condition_formset,
        'page_title': f'Create Rule for {custom_category.name}',
        'step': 2,
        'step_title': 'Step 2: Create Category Rule'
    }
    return render(request, 'analyzer/create_custom_category_rule.html', context)


@login_required
def custom_categories_list(request):
    """List all custom categories for the user"""
    custom_categories = CustomCategory.objects.filter(user=request.user).prefetch_related('rules')
    
    context = {
        'custom_categories': custom_categories,
        'page_title': 'My Custom Categories'
    }
    return render(request, 'analyzer/custom_categories_list.html', context)


@login_required
def edit_custom_category_rule(request, rule_id):
    """Edit a custom category rule"""
    rule = get_object_or_404(CustomCategoryRule, id=rule_id, user=request.user)
    
    if request.method == 'POST':
        rule_form = CustomCategoryRuleForm(request.POST, instance=rule)
        condition_formset = CustomCategoryRuleConditionFormSet(request.POST, instance=rule)
        
        if rule_form.is_valid() and condition_formset.is_valid():
            with db_transaction.atomic():
                rule_form.save()
                condition_formset.save()
                
                messages.success(request, f'Rule "{rule.name}" updated successfully!')
                return redirect('custom_categories_list')
    else:
        rule_form = CustomCategoryRuleForm(instance=rule)
        condition_formset = CustomCategoryRuleConditionFormSet(instance=rule)
    
    context = {
        'custom_category': rule.custom_category,
        'rule': rule,
        'rule_form': rule_form,
        'condition_formset': condition_formset,
        'page_title': f'Edit Rule: {rule.name}',
        'is_edit': True
    }
    return render(request, 'analyzer/create_custom_category_rule.html', context)


@login_required
def delete_custom_category_rule(request, rule_id):
    """Delete a custom category rule"""
    rule = get_object_or_404(CustomCategoryRule, id=rule_id, user=request.user)
    category_name = rule.custom_category.name
    
    if request.method == 'POST':
        rule.delete()
        messages.success(request, f'Rule deleted successfully!')
        return redirect('custom_categories_list')
    
    context = {
        'rule': rule,
        'page_title': 'Delete Rule'
    }
    return render(request, 'analyzer/delete_custom_category_rule.html', context)


@login_required
def delete_custom_category(request, category_id):
    """Delete a custom category"""
    custom_category = get_object_or_404(CustomCategory, id=category_id, user=request.user)
    
    if request.method == 'POST':
        category_name = custom_category.name
        custom_category.delete()
        messages.success(request, f'Category "{category_name}" deleted successfully!')
        return redirect('custom_categories_list')
    
    context = {
        'custom_category': custom_category,
        'page_title': 'Delete Custom Category'
    }
    return render(request, 'analyzer/delete_custom_category.html', context)


@login_required
def apply_custom_category(request, statement_id):
    """Apply multiple custom categories to transactions in a statement"""
    if request.method == 'POST':
        category_ids = request.POST.getlist('category_ids')
        
        if not category_ids:
            return JsonResponse({
                'success': False,
                'message': 'Please select at least one category to apply.',
                'matched_transaction_ids': []
            })
        
        try:
            statement = get_object_or_404(BankStatement, id=statement_id, account__user=request.user)
            
            # Get all selected custom categories
            custom_categories = CustomCategory.objects.filter(
                id__in=category_ids,
                user=request.user
            )
            
            if not custom_categories.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'No categories found.',
                    'matched_transaction_ids': []
                })
            
            # Apply rules to transactions
            from .rules_engine import CustomCategoryRulesEngine
            engine = CustomCategoryRulesEngine(request.user)
            
            transactions = Transaction.objects.filter(statement=statement)
            matched_transaction_ids = []
            category_names = []
            category_colors = []
            
            for category in custom_categories:
                category_names.append(category.name)
                category_colors.append(category.color)
                
                # Get the rules for this custom category
                rules = CustomCategoryRule.objects.filter(custom_category=category, is_active=True)
                
                if not rules.exists():
                    continue
                
                for transaction in transactions:
                    transaction_data = {
                        'description': transaction.description,
                        'amount': transaction.amount,
                        'date': transaction.date
                    }
                    
                    # Check if transaction matches the custom category rules
                    matching_category = engine.apply_rules_to_transaction(transaction_data)
                    
                    # Only apply if the matching category is our selected category
                    if matching_category and matching_category.id == category.id:
                        if transaction.id not in matched_transaction_ids:
                            matched_transaction_ids.append(transaction.id)
            
            if matched_transaction_ids:
                return JsonResponse({
                    'success': True,
                    'message': f'Applied {len(custom_categories)} categor{"ies" if len(custom_categories) > 1 else "y"} to {len(matched_transaction_ids)} matching transactions.',
                    'category_names': category_names,
                    'category_colors': category_colors,
                    'matched_transaction_ids': matched_transaction_ids,
                    'applied_count': len(matched_transaction_ids)
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': f'No transactions matched the rules for the selected categor{"ies" if len(custom_categories) > 1 else "y"}.',
                    'matched_transaction_ids': []
                })
        
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}',
                'matched_transaction_ids': []
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.',
        'matched_transaction_ids': []
    })


@login_required
def apply_custom_category_rules(request):
    """Apply multiple custom categories to transactions from rule application results"""
    if request.method == 'POST':
        category_ids = request.POST.getlist('category_ids')
        
        if not category_ids:
            return JsonResponse({
                'success': False,
                'message': 'Please select at least one category to apply.',
                'matched_transaction_ids': []
            })
        
        try:
            # Get all selected custom categories with their rules and conditions
            custom_categories = CustomCategory.objects.filter(
                id__in=category_ids,
                user=request.user
            ).prefetch_related('rules__conditions')
            
            if not custom_categories.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'No categories found for your account.',
                    'matched_transaction_ids': []
                })
            
            # Import the engine
            from .rules_engine import CustomCategoryRulesEngine
            
            # Get all transactions for this user
            transactions = Transaction.objects.filter(
                statement__account__user=request.user
            )
            
            if not transactions.exists():
                return JsonResponse({
                    'success': False,
                    'message': 'No transactions found in your account.',
                    'matched_transaction_ids': []
                })
            
            matched_transaction_ids = []
            category_names = []
            category_colors = []
            
            for category in custom_categories:
                category_names.append(category.name)
                category_colors.append(category.color)
                
                # Get the active rules for this custom category (from prefetched data)
                active_rules = [rule for rule in category.rules.all() if rule.is_active]
                
                if not active_rules:
                    # No active rules for this category, skip it
                    continue
                
                for transaction in transactions:
                    transaction_data = {
                        'description': transaction.description,
                        'amount': float(transaction.amount),
                        'date': transaction.date
                    }
                    
                    # Check if transaction matches ANY of this category's rules
                    matches_category = False
                    for rule in active_rules:
                        # Get all conditions for this rule
                        conditions = list(rule.conditions.all())
                        
                        if not conditions:
                            # If rule has no conditions, skip it
                            continue
                        
                        if CustomCategoryRulesEngine._matches_rule_static(transaction_data, rule):
                            matches_category = True
                            break
                    
                    if matches_category:
                        if transaction.id not in matched_transaction_ids:
                            matched_transaction_ids.append(transaction.id)
            
            if matched_transaction_ids:
                return JsonResponse({
                    'success': True,
                    'message': f'Applied {len(custom_categories)} categor{"ies" if len(custom_categories) > 1 else "y"} to {len(matched_transaction_ids)} matching transactions.',
                    'category_names': category_names,
                    'category_colors': category_colors,
                    'matched_transaction_ids': matched_transaction_ids,
                    'applied_count': len(matched_transaction_ids)
                })
            else:
                # Check if we have any categories with active rules
                categories_with_rules = [cat for cat in custom_categories if any(r.is_active for r in cat.rules.all())]
                
                if not categories_with_rules:
                    return JsonResponse({
                        'success': False,
                        'message': 'Selected categories have no active rules. Please activate the rules first.',
                        'matched_transaction_ids': []
                    })
                
                return JsonResponse({
                    'success': False,
                    'message': f'No transactions matched the rules for the selected categor{"ies" if len(custom_categories) > 1 else "y"}.',
                    'matched_transaction_ids': []
                })
        
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.exception(f"Error in apply_custom_category_rules: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}',
                'matched_transaction_ids': []
            })
    
    return JsonResponse({
        'success': False,
        'message': 'Invalid request method.',
        'matched_transaction_ids': []
    })


@login_required
def get_financial_overview_data(request):
    """Get financial overview data for different time periods (AJAX endpoint)"""
    time_period = request.GET.get('period', 'all')
    
    # Get all transactions for this user
    all_transactions = Transaction.objects.filter(
        statement__account__user=request.user
    )
    
    # Filter by time period
    now = timezone.now().date()
    if time_period == '30days':
        start_date = now - timedelta(days=30)
        transactions = all_transactions.filter(date__gte=start_date)
    elif time_period == '90days':
        start_date = now - timedelta(days=90)
        transactions = all_transactions.filter(date__gte=start_date)
    else:  # all time
        transactions = all_transactions
    
    # Calculate income and expenses
    total_income = transactions.filter(transaction_type='CREDIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    total_expenses = transactions.filter(transaction_type='DEBIT').aggregate(
        total=models.Sum('amount')
    )['total'] or 0
    
    net_savings = total_income - total_expenses
    
    # Calculate percentages
    total_all = total_income + total_expenses
    income_percentage = (total_income / total_all * 100) if total_all > 0 else 0
    expense_percentage = (total_expenses / total_all * 100) if total_all > 0 else 0
    
    # Calculate financial health
    if total_income > 0:
        savings_rate = (net_savings / total_income) * 100
        if savings_rate >= 20:
            health_status = 'Excellent'
            health_score = 85
        elif savings_rate >= 10:
            health_status = 'Good'
            health_score = 70
        else:
            health_status = 'Needs Attention'
            health_score = 50
    else:
        health_status = 'No Data'
        health_score = 0
    
    return JsonResponse({
        'success': True,
        'income': float(total_income),
        'expenses': float(total_expenses),
        'savings': float(net_savings),
        'income_percentage': round(income_percentage, 1),
        'expense_percentage': round(expense_percentage, 1),
        'health_status': health_status,
        'health_score': health_score,
        'transaction_count': transactions.count(),
        'period': time_period
    })


@login_required
@login_required
def export_rules_results_to_excel(request):
    """Export rule application results to Excel file with filtered transactions only"""
    
    # Check if openpyxl is available
    if not EXCEL_EXPORT_AVAILABLE:
        print("ERROR - openpyxl is not available")
        messages.error(request, 'Excel export not available. Please install openpyxl.')
        return redirect('rules_application_results')
    
    try:
        from datetime import datetime
        from django.db.models import Sum
        
        print("DEBUG - Starting export process")
        
        # Get filtered results from session
        export_filtered_results = request.session.get('export_filtered_results', [])
        selected_rule_ids_str = request.session.get('export_selected_rule_ids', [])
        selected_category_ids_str = request.session.get('export_selected_category_ids', [])
        
        # Convert to integers
        selected_rule_ids = [int(rid) for rid in selected_rule_ids_str if rid.isdigit()] if isinstance(selected_rule_ids_str, list) else []
        selected_category_ids = [int(cid) for cid in selected_category_ids_str if cid.isdigit()] if isinstance(selected_category_ids_str, list) else []
        
        # Also get from POST for backup (if called directly)
        if not export_filtered_results:
            selected_rule_ids = request.POST.getlist('rule_ids') if request.method == 'POST' else []
            selected_category_ids = request.POST.getlist('category_ids') if request.method == 'POST' else []
            selected_rule_ids = [int(rid) for rid in selected_rule_ids if rid.isdigit()]
            selected_category_ids = [int(cid) for cid in selected_category_ids if cid.isdigit()]
        
        print(f"DEBUG - Filtered Results Count: {len(export_filtered_results)}")
        print(f"DEBUG - Selected Rule IDs: {selected_rule_ids}")
        print(f"DEBUG - Selected Category IDs: {selected_category_ids}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Filtered Results'
        
        # Define styles
        header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        section_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        summary_fill = PatternFill(start_color='E8F0F7', end_color='E8F0F7', fill_type='solid')
        total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Professional Report Header Layout
        current_date = datetime.now()
        current_row = 1
        
        # Row 1: Date and Day on left side
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"Date: {current_date.strftime('%Y-%m-%d')}   Day: {current_date.strftime('%A')}"
        cell.font = Font(bold=True, size=10)
        current_row += 1
        
        # Row 2: Empty line for spacing
        current_row += 1
        
        # Row 3: Project Name "BANKWATCH" - Centered, merged across columns
        ws.merge_cells(f'A{current_row}:G{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "BANKWATCH - Filtered Transactions Report"
        cell.font = Font(bold=True, size=14, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Row 4: Selected Rules
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Selected Rules:"
        cell.font = Font(bold=True, size=11)
        
        try:
            selected_rules = Rule.objects.filter(id__in=selected_rule_ids, user=request.user)
            rule_names = ', '.join([rule.name for rule in selected_rules]) if selected_rules.exists() else "None"
        except:
            rule_names = "None"
        
        cell = ws.cell(row=current_row, column=2)
        cell.value = rule_names
        current_row += 1
        
        # Row 5: Selected Categories
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Selected Categories:"
        cell.font = Font(bold=True, size=11)
        
        try:
            selected_categories = CustomCategory.objects.filter(id__in=selected_category_ids, user=request.user)
            category_names = ', '.join([cat.name for cat in selected_categories]) if selected_categories.exists() else "None"
        except:
            category_names = "None"
        
        cell = ws.cell(row=current_row, column=2)
        cell.value = category_names
        current_row += 2
        
        # Prepare headers - include rule and category columns
        headers = ['Date', 'Account', 'Description', 'Amount', 'Matched Rule', 'Category Applied']
        column_widths = [12, 18, 35, 12, 20, 20]
        
        print(f"DEBUG - Headers: {headers}")
        
        # Add headers
        table_start_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Set column widths
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width
        
        print(f"DEBUG - Starting to add {len(export_filtered_results)} filtered transactions")
        
        # Add only filtered transaction data
        row_num = table_start_row + 1
        total_amount = 0
        
        for result in export_filtered_results:
            try:
                # Fetch the actual transaction object
                tx = Transaction.objects.get(id=result['id'], statement__account__user=request.user)
                
                col_num = 1
                
                # Date
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = tx.date.strftime('%Y-%m-%d') if tx.date else ''
                cell.border = border
                cell.alignment = center_align
                col_num += 1
                
                # Account
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = result['account_name']
                cell.border = border
                cell.alignment = left_align
                col_num += 1
                
                # Description
                desc = result['description'] if result['description'] else ''
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = desc
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                col_num += 1
                
                # Amount
                amount = float(result['amount']) if result['amount'] else 0
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = amount
                cell.border = border
                cell.number_format = '#,##0.00'
                cell.alignment = center_align
                total_amount += amount
                col_num += 1
                
                # Matched Rule Name
                matched_rule_name = result.get('matched_rule_name', '')
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = matched_rule_name if matched_rule_name else '-'
                cell.border = border
                cell.alignment = left_align
                col_num += 1
                
                # Custom Category Name
                matched_category_name = result.get('matched_custom_category_name', '')
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = matched_category_name if matched_category_name else '-'
                cell.border = border
                cell.alignment = left_align
                
                row_num += 1
            except Exception as e:
                print(f"ERROR - Failed to process result {result.get('id')}: {e}")
                continue
        
        print(f"DEBUG - Completed adding transactions at row {row_num}")
        
        # Add Summary Section
        summary_row = row_num + 2
        
        # Summary Header
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "FILTERED SUMMARY"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        cell.border = border
        summary_row += 1
        
        # Total Transactions (from filtered results only)
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Filtered Transactions:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = len(export_filtered_results)
        cell.font = Font(size=11)
        cell.border = border
        summary_row += 1
        
        # Total Amount (from filtered results only)
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Filtered Amount:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = total_amount
        cell.font = Font(size=11)
        cell.border = border
        cell.number_format = '#,##0.00'
        summary_row += 1
        
        # Rules Selected Count
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Rules Selected:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = len(selected_rule_ids)
        cell.font = Font(size=11)
        cell.border = border
        summary_row += 1
        
        # Categories Selected Count
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Categories Selected:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = len(selected_category_ids)
        cell.font = Font(size=11)
        cell.border = border
        
        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=filtered_rules_results.xlsx'
        
        wb.save(response)
        return response
        
    except Exception as e:
        error_msg = f"Exception in export_rules_results_to_excel: {str(e)}"
        print(f"ERROR - {error_msg}")
        messages.error(request, f'Error exporting to Excel: {error_msg}')
        return redirect('rules_application_results')
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Results'
        
        # Define styles
        header_fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        section_fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
        summary_fill = PatternFill(start_color='E8F0F7', end_color='E8F0F7', fill_type='solid')
        total_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        # Build summary data
        try:
            selected_rules = Rule.objects.filter(id__in=selected_rule_ids, user=request.user).order_by('name')
            selected_categories = CustomCategory.objects.filter(id__in=selected_category_ids, user=request.user).order_by('name')
        except Exception as e:
            print(f"ERROR - Failed to fetch rules/categories: {e}")
            selected_rules = Rule.objects.none()
            selected_categories = CustomCategory.objects.none()
        
        rule_totals_dict = {}
        category_totals_dict = {}
        
        # Calculate rule totals
        for rule in selected_rules:
            try:
                rule_count = 0
                rule_total = 0
                for tx in all_transactions:
                    if tx.category == rule.category:
                        rule_count += 1
                        rule_total += float(tx.amount)
                rule_totals_dict[rule.name] = {'total': rule_total, 'count': rule_count}
            except Exception as e:
                print(f"ERROR - Failed to calculate rule total for {rule.name}: {e}")
        
        # Calculate category totals (based on matching custom category rules)
        for category in selected_categories:
            try:
                category_count = 0
                category_total = 0
                rules = CustomCategoryRule.objects.filter(custom_category=category, user=request.user)
                
                for tx in all_transactions:
                    # Check if transaction matches this category's rules
                    if rules.exists():
                        for rule in rules:
                            conditions = rule.conditions.all()
                            if conditions.exists():
                                match = True
                                for condition in conditions:
                                    field_value = getattr(tx, condition.field, '')
                                    if condition.condition_type == 'contains':
                                        if condition.value.lower() not in str(field_value).lower():
                                            match = False
                                            break
                                    elif condition.condition_type == 'exact':
                                        if condition.value.lower() != str(field_value).lower():
                                            match = False
                                            break
                                if match:
                                    category_count += 1
                                    category_total += float(tx.amount) if tx.amount else 0
                                    break
                
                if category_count > 0 or category_total > 0:
                    category_totals_dict[category.name] = {'total': category_total, 'count': category_count}
            except Exception as e:
                print(f"ERROR - Failed to calculate category total for {category.name}: {e}")
        
        print("DEBUG - Starting to build Excel sheet")
        
        # Professional Report Header Layout
        current_date = datetime.now()
        current_row = 1
        
        # Row 1: Date and Day on left side
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"Date: {current_date.strftime('%Y-%m-%d')}   Day: {current_date.strftime('%A')}"
        cell.font = Font(bold=True, size=10)
        current_row += 1
        
        # Row 2: Empty line for spacing
        current_row += 1
        
        # Row 3: Project Name "BANKWATCH" - Centered, merged across columns
        ws.merge_cells(f'A{current_row}:H{current_row}')
        cell = ws.cell(row=current_row, column=1)
        cell.value = "BANKWATCH"
        cell.font = Font(bold=True, size=16, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # Row 4: Selected Rules
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Selected Rules:"
        cell.font = Font(bold=True, size=11)
        
        if selected_rules.exists():
            rule_names = ', '.join([rule.name for rule in selected_rules])
            cell = ws.cell(row=current_row, column=2)
            cell.value = rule_names
        else:
            cell = ws.cell(row=current_row, column=2)
            cell.value = "NULL"
            cell.font = Font(bold=True, size=11)
        
        current_row += 1
        
        # Row 5: Selected Categories
        cell = ws.cell(row=current_row, column=1)
        cell.value = "Selected Categories:"
        cell.font = Font(bold=True, size=11)
        
        if selected_categories.exists():
            category_names = ', '.join([cat.name for cat in selected_categories])
            cell = ws.cell(row=current_row, column=2)
            cell.value = category_names
        else:
            cell = ws.cell(row=current_row, column=2)
            cell.value = "NULL"
            cell.font = Font(bold=True, size=11)
        
        # Row for results table header
        current_row += 2
        
        # Prepare headers based on whether show_changed is enabled
        if show_changed:
            headers = ['Date', 'Account', 'Description', 'Amount', 'Previous Category', 'Current Category', 'Matched Rule', 'Custom Category']
            column_widths = [12, 18, 40, 12, 18, 18, 20, 18]
        else:
            headers = ['Date', 'Account', 'Description', 'Amount', 'Category', 'Matched Rule', 'Custom Category']
            column_widths = [12, 18, 40, 12, 18, 20, 18]
        
        print(f"DEBUG - Headers: {headers}")
        print(f"DEBUG - Column widths: {column_widths}")
        
        # Add headers
        table_start_row = current_row
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=table_start_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        print(f"DEBUG - Starting to add {all_transactions.count()} transactions")
        
        # Add transaction data
        row_num = table_start_row + 1
        for transaction in all_transactions:
            try:
                col_num = 1
                
                # Date
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = transaction.date.strftime('%Y-%m-%d') if transaction.date else ''
                cell.border = border
                cell.alignment = center_align
                col_num += 1
                
                # Account
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = transaction.statement.account.account_name if transaction.statement else ''
                cell.border = border
                cell.alignment = left_align
                col_num += 1
                
                # Description - Truncate to 45 characters to prevent merging
                desc = transaction.description if transaction.description else ''
                if len(desc) > 45:
                    desc = desc[:42] + '...'
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = desc
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                ws.row_dimensions[row_num].height = None  # Auto-adjust row height
                col_num += 1
                
                # Amount
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = float(transaction.amount) if transaction.amount else 0
                cell.border = border
                cell.number_format = '#,##0.00'
                cell.alignment = center_align
                col_num += 1
                
                if show_changed:
                    # Previous Category (from session data)
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = '-'
                    cell.border = border
                    cell.alignment = left_align
                    col_num += 1
                
                # Current Category
                cell = ws.cell(row=row_num, column=col_num)
                try:
                    category_display = transaction.get_category_display()
                except:
                    category_display = str(transaction.category)
                cell.value = category_display if category_display else '-'
                cell.border = border
                cell.alignment = left_align
                col_num += 1
                
                # Matched Rule - simplified
                matched_rule = '-'
                try:
                    # Check if this transaction matches any selected rule
                    if selected_rules.exists():
                        for rule in selected_rules:
                            if rule.category == transaction.category:
                                matched_rule = rule.name
                                break
                except Exception as e:
                    print(f"ERROR - Failed to match rule: {e}")
                
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = matched_rule
                cell.border = border
                cell.alignment = left_align
                col_num += 1
                
                # Custom Category
                cell = ws.cell(row=row_num, column=col_num)
                custom_cat = '-'
                # Check if transaction matches any selected custom category
                try:
                    if selected_categories.exists():
                        for category in selected_categories:
                            # Check if transaction matches this category's rules
                            rules = CustomCategoryRule.objects.filter(custom_category=category, user=request.user)
                            for rule in rules:
                                # Simple match: check if description contains the rule keyword
                                conditions = rule.conditions.all()
                                if conditions.exists():
                                    match = True
                                    for condition in conditions:
                                        field_value = getattr(transaction, condition.field, '')
                                        if condition.condition_type == 'contains':
                                            if condition.value.lower() not in str(field_value).lower():
                                                match = False
                                                break
                                        elif condition.condition_type == 'exact':
                                            if condition.value.lower() != str(field_value).lower():
                                                match = False
                                                break
                                    if match:
                                        custom_cat = category.name
                                        break
                except Exception as e:
                    print(f"DEBUG - Custom category match error: {e}")
                    
                cell.value = custom_cat
                cell.border = border
                cell.alignment = left_align
                
                row_num += 1
            except Exception as e:
                print(f"ERROR - Failed to process transaction {transaction.id}: {e}")
                continue
        
        print(f"DEBUG - Completed adding transactions at row {row_num}")
        
        # Add Summary Section after the table
        summary_row = row_num + 2
        
        # Summary Header
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "FINAL SUMMARY"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='0D47A1', end_color='0D47A1', fill_type='solid')
        cell.border = border
        summary_row += 1
        
        # Total Transactions
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Transactions:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = all_transactions.count()
        cell.font = Font(size=11)
        cell.border = border
        summary_row += 1
        
        # Total Transaction Amount
        total_transaction_amount = sum(float(tx.amount) for tx in all_transactions if tx.amount)
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Transaction Amount:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = f"₹{total_transaction_amount:,.2f}"
        cell.font = Font(size=11)
        cell.border = border
        cell.number_format = '₹#,##0.00'
        summary_row += 1
        
        # Total Rules Selected
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Rules Selected:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = selected_rules.count()
        cell.font = Font(size=11)
        cell.border = border
        summary_row += 1
        
        # Total Categories Selected
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "Total Categories Selected:"
        cell.font = Font(bold=True, size=11)
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = selected_categories.count()
        cell.font = Font(size=11)
        cell.border = border
        summary_row += 1
        
        # Total Rule Amount (only if rules selected)
        total_rules_amount = sum(item['total'] for item in rule_totals_dict.values())
        if total_rules_amount > 0:
            cell = ws.cell(row=summary_row, column=1)
            cell.value = "Total Rule Amount:"
            cell.font = Font(bold=True, size=11)
            cell.border = border
            
            cell = ws.cell(row=summary_row, column=2)
            cell.value = f"₹{total_rules_amount:,.2f}"
            cell.font = Font(size=11)
            cell.border = border
            cell.number_format = '₹#,##0.00'
            summary_row += 1
        
        # Total Category Amount (only if categories selected)
        total_categories_amount = sum(item['total'] for item in category_totals_dict.values())
        if total_categories_amount > 0:
            cell = ws.cell(row=summary_row, column=1)
            cell.value = "Total Category Amount:"
            cell.font = Font(bold=True, size=11)
            cell.border = border
            
            cell = ws.cell(row=summary_row, column=2)
            cell.value = f"₹{total_categories_amount:,.2f}"
            cell.font = Font(size=11)
            cell.border = border
            cell.number_format = '₹#,##0.00'
            summary_row += 1
        
        # GRAND TOTAL AMOUNT
        summary_row += 1
        grand_total = total_rules_amount + total_categories_amount
        cell = ws.cell(row=summary_row, column=1)
        cell.value = "GRAND TOTAL (Rules + Categories):"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        cell.border = border
        
        cell = ws.cell(row=summary_row, column=2)
        cell.value = f"₹{grand_total:,.2f}"
        cell.font = Font(bold=True, size=12, color='FFFFFF')
        cell.fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        cell.border = border
        cell.number_format = '₹#,##0.00'
        
        # Adjust column widths
        for i, width in enumerate(column_widths, 1):
            col_letter = chr(64 + i)  # Convert to A, B, C, etc.
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row of data table
        ws.freeze_panes = f'A{table_start_row + 1}'
        
        print("DEBUG - Creating HTTP response")
        
        # Create HTTP response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="rule_results.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        print("DEBUG - Excel file generated successfully")
        return response
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        print(f"ERROR - Exception in export_rules_results_to_excel: {error_msg}")
        print(f"ERROR - Traceback: {traceback_str}")
        messages.error(request, f'Error exporting results: {error_msg}')
        return redirect('rules_application_results')

@login_required
def export_rules_results_to_pdf(request):
    """Export filtered rule application results to PDF file"""
    
    try:
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.colors import HexColor, black, white, lightgrey
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from io import BytesIO
        from datetime import datetime
        
        print("DEBUG - Starting PDF export process")
        
        # Get filtered results from session
        export_filtered_results = request.session.get('export_filtered_results', [])
        selected_rule_ids_str = request.session.get('export_selected_rule_ids', [])
        selected_category_ids_str = request.session.get('export_selected_category_ids', [])
        
        # Convert to integers
        selected_rule_ids = [int(rid) for rid in selected_rule_ids_str if rid.isdigit()] if isinstance(selected_rule_ids_str, list) else []
        selected_category_ids = [int(cid) for cid in selected_category_ids_str if cid.isdigit()] if isinstance(selected_category_ids_str, list) else []
        
        print(f"DEBUG - Filtered Results Count: {len(export_filtered_results)}")
        print(f"DEBUG - Selected Rule IDs: {selected_rule_ids}")
        print(f"DEBUG - Selected Category IDs: {selected_category_ids}")
        
        # Create PDF document in memory using landscape orientation for better column layout
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=landscape(letter),
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Container for PDF elements
        elements = []
        
        # Define styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=white,
            backColor=HexColor('0D47A1'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=11,
            textColor=HexColor('0D47A1'),
            spaceAfter=6,
            fontName='Helvetica-Bold'
        )
        
        normal_style = styles['Normal']
        
        # Add title
        elements.append(Paragraph("BANKWATCH - Filtered Transactions Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Add report metadata
        current_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elements.append(Paragraph(f"Report Generated: {current_date}", normal_style))
        
        # Get rule and category names
        try:
            selected_rules = Rule.objects.filter(id__in=selected_rule_ids, user=request.user)
            rule_names = ', '.join([rule.name for rule in selected_rules]) if selected_rules.exists() else "None"
        except:
            rule_names = "None"
        
        try:
            selected_categories = CustomCategory.objects.filter(id__in=selected_category_ids, user=request.user)
            category_names = ', '.join([cat.name for cat in selected_categories]) if selected_categories.exists() else "None"
        except:
            category_names = "None"
        
        elements.append(Paragraph(f"<b>Selected Rules:</b> {rule_names}", normal_style))
        elements.append(Paragraph(f"<b>Selected Categories:</b> {category_names}", normal_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Build transaction table with proper column widths
        # Column widths: Date, Account, Description (wider), Amount, Rule, Category
        # Adjust column widths for better description display: Date, Account, Description (wider), Amount, Rule, Category
        col_widths = [0.9*inch, 0.9*inch, 3.2*inch, 0.85*inch, 1.2*inch, 1.2*inch]
        
        # Table header
        table_data = [
            ['Date', 'Account', 'Description', 'Amount', 'Matched Rule', 'Category Applied']
        ]
        
        total_amount = 0
        transaction_count = 0
        
        # Create a style for wrapped text in description
        desc_style = ParagraphStyle(
            'DescriptionStyle',
            parent=normal_style,
            fontSize=8,
            alignment=TA_LEFT,
            wordWrap='CJK',
            leading=10
        )
        
        # Add transaction rows from filtered results
        for result in export_filtered_results:
            try:
                date_str = result['date'].strftime('%Y-%m-%d') if hasattr(result['date'], 'strftime') else str(result['date'])
                amount = float(result['amount']) if result['amount'] else 0
                # Use full description text - wrap it in Paragraph for proper handling
                description = result['description'] if result['description'] else ''
                # Create Paragraph object for proper text wrapping in table cell
                description_para = Paragraph(description, desc_style)
                
                matched_rule = result.get('matched_rule_name', '') or '-'
                matched_category = result.get('matched_custom_category_name', '') or '-'
                
                table_data.append([
                    date_str,
                    result['account_name'],
                    description_para,  # Use Paragraph for wrapping
                    f"₹{amount:,.2f}",
                    matched_rule,
                    matched_category
                ])
                
                total_amount += amount
                transaction_count += 1
            except Exception as e:
                print(f"ERROR - Failed to process result: {e}")
                continue
        
        # Add summary row
        table_data.append([
            '', '', Paragraph('<b>TOTAL</b>', desc_style),
            Paragraph(f'<b>₹{total_amount:,.2f}</b>', desc_style),
            '', ''
        ])
        
        # Create table with proper styling
        table = Table(table_data, colWidths=col_widths, splitByRow=True)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -2), 1, lightgrey),
            
            # Data rows
            ('ALIGN', (0, 1), (-1, -2), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -2), 'TOP'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [white, HexColor('F5F5F5')]),
            
            # Description column - left align, allow wrapping
            ('ALIGN', (2, 0), (2, -1), 'LEFT'),
            ('VALIGN', (2, 0), (2, -1), 'TOP'),
            
            # Amount column - right align
            ('ALIGN', (3, 1), (3, -2), 'RIGHT'),
            ('ALIGN', (3, -1), (3, -1), 'RIGHT'),
            
            # Rule and Category columns - left align
            ('ALIGN', (4, 1), (-1, -2), 'LEFT'),
            
            # Total row
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('FFF2CC')),
            ('ALIGN', (2, -1), (2, -1), 'LEFT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, -1), (-1, -1), 8),
            ('BOTTOMPADDING', (0, -1), (-1, -1), 8),
            
            # Padding for all cells
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 1), (-1, -2), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -2), 6),
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Add summary section
        elements.append(Paragraph("FILTERED SUMMARY", heading_style))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Total Filtered Transactions', str(transaction_count)],
            ['Total Filtered Amount', f'₹{total_amount:,.2f}'],
            ['Rules Selected', str(len(selected_rule_ids))],
            ['Categories Selected', str(len(selected_category_ids))],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, lightgrey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('F5F5F5')]),
        ]))
        
        elements.append(summary_table)
        
        # Build PDF
        doc.build(elements)
        
        # Prepare response
        pdf_buffer.seek(0)
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename=filtered_rules_results.pdf'
        
        return response
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        print(f"ERROR - Exception in export_rules_results_to_pdf: {error_msg}")
        print(f"ERROR - Traceback: {traceback.format_exc()}")
        messages.error(request, f'Error exporting to PDF: {error_msg}')
        return redirect('rules_application_results')

        
        rule_totals_dict = {}
        category_totals_dict = {}
        category_chart_data = {}
        
        # Calculate rule totals
        for rule in selected_rules:
            try:
                rule_count = 0
                rule_total = 0
                for tx in all_transactions:
                    if tx.category == rule.category:
                        rule_count += 1
                        rule_total += float(tx.amount)
                rule_totals_dict[rule.name] = {'total': rule_total, 'count': rule_count}
            except Exception as e:
                print(f"ERROR - Failed to calculate rule total for {rule.name}: {e}")
        
        # Calculate category totals for chart
        expense_transactions = all_transactions.filter(transaction_type='DEBIT')
        for category_code, category_name in Transaction.CATEGORY_CHOICES:
            if category_code != 'INCOME':
                category_total = expense_transactions.filter(category=category_code).aggregate(
                    total=models.Sum('amount')
                )['total'] or 0
                if category_total > 0:
                    category_chart_data[category_name] = float(category_total)
        
        # Calculate category totals (based on matching custom category rules)
        for category in selected_categories:
            try:
                category_count = 0
                category_total = 0
                rules = CustomCategoryRule.objects.filter(custom_category=category, user=request.user)
                
                for tx in all_transactions:
                    # Check if transaction matches this category's rules
                    if rules.exists():
                        for rule in rules:
                            conditions = rule.conditions.all()
                            if conditions.exists():
                                match = True
                                for condition in conditions:
                                    field_value = getattr(tx, condition.field, '')
                                    if condition.condition_type == 'contains':
                                        if condition.value.lower() not in str(field_value).lower():
                                            match = False
                                            break
                                    elif condition.condition_type == 'exact':
                                        if condition.value.lower() != str(field_value).lower():
                                            match = False
                                            break
                                if match:
                                    category_count += 1
                                    category_total += float(tx.amount) if tx.amount else 0
                                    break
                
                if category_count > 0 or category_total > 0:
                    category_totals_dict[category.name] = {'total': category_total, 'count': category_count}
            except Exception as e:
                print(f"ERROR - Failed to calculate category total for {category.name}: {e}")
        
        print("DEBUG - Starting to build PDF document")
        
        # Create PDF
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
        elements = []
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=HexColor('#0D47A1'),
            spaceAfter=10,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        header_style = ParagraphStyle(
            'CustomHeader',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=HexColor('#0D47A1'),
            spaceAfter=8,
            spaceBefore=8,
            fontName='Helvetica-Bold'
        )
        
        # Title
        title = Paragraph("BANKWATCH - Financial Analysis Report", title_style)
        elements.append(title)
        
        # Date and metadata
        current_date = datetime.now()
        date_text = f"<b>Date:</b> {current_date.strftime('%Y-%m-%d')} | <b>Day:</b> {current_date.strftime('%A')}"
        elements.append(Paragraph(date_text, styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Selected Rules and Categories
        if selected_rules.exists():
            rule_names = ', '.join([rule.name for rule in selected_rules])
            elements.append(Paragraph(f"<b>Selected Rules:</b> {rule_names}", styles['Normal']))
        
        if selected_categories.exists():
            category_names = ', '.join([cat.name for cat in selected_categories])
            elements.append(Paragraph(f"<b>Selected Categories:</b> {category_names}", styles['Normal']))
        
        elements.append(Spacer(1, 0.3*inch))
        
        # Create pie chart from category data
        if category_chart_data:
            try:
                import matplotlib
                import matplotlib.pyplot as plt
                matplotlib.use('Agg')  # Use non-GUI backend before creating figure
                
                fig, ax = plt.subplots(figsize=(8, 6))
                categories = list(category_chart_data.keys())
                amounts = list(category_chart_data.values())
                
                colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#FFA07A', '#98D8C8', '#F7DC6F', '#BB8FCE', '#85C1E2']
                ax.pie(amounts, labels=categories, autopct='%1.1f%%', startangle=90, colors=colors[:len(categories)])
                ax.set_title('Spending by Category', fontsize=14, fontweight='bold')
                
                # Save chart to temporary file
                import os
                chart_file = os.path.join(tempfile.gettempdir(), f'chart_{int(datetime.now().timestamp())}.png')
                plt.savefig(chart_file, format='png', bbox_inches='tight', dpi=100)
                plt.close(fig)
                
                # Add chart to PDF
                from reportlab.platypus import Image as ReportLabImage
                if os.path.exists(chart_file):
                    chart_image = ReportLabImage(chart_file, width=5*inch, height=4*inch)
                    elements.append(chart_image)
                    elements.append(Spacer(1, 0.2*inch))
                    
                    # Clean up temp file after building PDF
                    atexit.register(lambda: os.unlink(chart_file) if os.path.exists(chart_file) else None)
                else:
                    print(f"ERROR - Chart file not created at {chart_file}")
                    elements.append(Paragraph("(Chart file not found)", styles['Normal']))
                    
            except Exception as e:
                import traceback
                print(f"ERROR - Failed to create chart: {e}")
                print(f"Traceback: {traceback.format_exc()}")
                elements.append(Paragraph(f"(Chart generation failed: {str(e)[:50]})", styles['Normal']))
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary Section
        elements.append(Paragraph("SUMMARY", header_style))
        
        # Create summary table
        summary_data = [['Metric', 'Value']]
        summary_data.append(['Total Transactions', str(all_transactions.count())])
        
        # Calculate total transaction amount
        total_transaction_amount = sum(float(tx.amount) for tx in all_transactions if tx.amount)
        summary_data.append(['Total Transaction Amount', f"₹{total_transaction_amount:,.2f}"])
        
        summary_data.append(['Total Rules Selected', str(selected_rules.count())])
        summary_data.append(['Total Categories Selected', str(selected_categories.count())])
        
        total_rules_amount = sum(item['total'] for item in rule_totals_dict.values())
        if total_rules_amount > 0:
            summary_data.append(['Total Rule Amount', f"₹{total_rules_amount:,.2f}"])
        
        total_categories_amount = sum(item['total'] for item in category_totals_dict.values())
        if total_categories_amount > 0:
            summary_data.append(['Total Category Amount', f"₹{total_categories_amount:,.2f}"])
        
        grand_total = total_rules_amount + total_categories_amount
        if grand_total > 0:
            summary_data.append(['GRAND TOTAL (Rules + Categories)', f"₹{grand_total:,.2f}"])
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0D47A1')),
            ('TEXTCOLOR', (0, 0), (-1, 0), white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#00B050')),
            ('TEXTCOLOR', (0, -1), (-1, -1), white),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Transactions Detail Section
        elements.append(PageBreak())
        elements.append(Paragraph("TRANSACTION DETAILS", header_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Build transactions table
        if all_transactions.count() > 0:
            # Prepare table data
            if show_changed:
                headers = ['Date', 'Account', 'Description', 'Amount', 'Previous', 'Current', 'Rule', 'Category']
                col_widths = [0.7*inch, 0.9*inch, 1.5*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch]
            else:
                headers = ['Date', 'Account', 'Description', 'Amount', 'Category', 'Rule']
                col_widths = [0.7*inch, 0.9*inch, 1.5*inch, 0.9*inch, 0.9*inch, 0.9*inch]
            
            table_data = [headers]
            
            for transaction in all_transactions:
                try:
                    row = []
                    row.append(transaction.date.strftime('%Y-%m-%d') if transaction.date else '')
                    row.append(transaction.statement.account.account_name if transaction.statement else '')
                    # Truncate description more aggressively for PDF
                    desc = transaction.description[:35] if transaction.description else ''
                    row.append(desc)
                    row.append(f"₹{float(transaction.amount):,.2f}" if transaction.amount else '')
                    
                    if show_changed:
                        row.append('-')  # Previous category
                    
                    try:
                        category_display = transaction.get_category_display()
                    except:
                        category_display = str(transaction.category)
                    row.append(category_display if category_display else '-')
                    
                    # Matched Rule
                    matched_rule = '-'
                    try:
                        if selected_rules.exists():
                            for rule in selected_rules:
                                if rule.category == transaction.category:
                                    matched_rule = rule.name
                                    break
                    except Exception as e:
                        print(f"ERROR - Failed to match rule: {e}")
                    
                    row.append(matched_rule)
                    
                    table_data.append(row)
                except Exception as e:
                    print(f"ERROR - Failed to process transaction {transaction.id}: {e}")
                    continue
            
            # Create transactions table
            transactions_table = Table(table_data, colWidths=col_widths)
            transactions_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#0D47A1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, lightgrey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [white, HexColor('#F5F5F5')]),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ]))
            elements.append(transactions_table)
        else:
            elements.append(Paragraph("No transactions found", styles['Normal']))
        
        print("DEBUG - Building PDF document")
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF content
        pdf_content = pdf_buffer.getvalue()
        pdf_buffer.close()
        
        # Create HTTP response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="rule_results.pdf"'
        response.write(pdf_content)
        
        print("DEBUG - PDF generated successfully")
        return response
        
    except ImportError as e:
        import traceback
        print(f"ERROR - Missing required library: {e}")
        print(f"ERROR - Full traceback:\n{traceback.format_exc()}")
        error_response = HttpResponse(
            f"Error: PDF export requires reportlab and matplotlib. Please install them.\n\nDetails: {str(e)}\n\nTraceback:\n{traceback.format_exc()}",
            content_type='text/plain',
            status=500
        )
        return error_response
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        print(f"ERROR - Exception in export_rules_results_to_pdf: {error_msg}")
        print(f"ERROR - Traceback: {traceback_str}")
        error_response = HttpResponse(
            f"Error generating PDF: {error_msg}\n\n{traceback_str}",
            content_type='text/plain',
            status=500
        )
        return error_response